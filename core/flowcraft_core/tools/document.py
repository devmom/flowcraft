"""Document processing tools - PDF, Word, Excel readers.

Optional dependencies:
    PyMuPDF (fitz) - PDF reading
    python-docx - Word reading
    openpyxl - Excel reading
"""
from __future__ import annotations

import logging
from pathlib import Path

from flowcraft_core.domain.enums import RiskLevel
from flowcraft_core.domain.schemas import ToolIntent
from flowcraft_core.tools.base import Tool, ToolDefinition, is_path_allowed, observation_from_output

logger = logging.getLogger(__name__)


class PdfReadTool(Tool):
    def __init__(self, allowed_paths: list[Path]) -> None:
        self.allowed_paths = allowed_paths
        self.definition = ToolDefinition(
            tool_name="document.pdf.read",
            display_name="Read PDF",
            description="Read and extract text from PDF files",
            category="document",
            risk_level=RiskLevel.LOW,
            permissions=["file.read"],
            requires_approval_by_default=False,
        )

    async def execute(self, intent: ToolIntent):
        path = Path(intent.input_payload.get("path", ""))
        if not is_path_allowed(path, self.allowed_paths):
            return observation_from_output(intent, "DENIED",
                f"You need permission to access this file. Ask the user: 'I need to read a file outside the workspace, grant access?'",
                error="Path not allowed.",
                payload={"action": "ask_user_for_permission"})
        if not path.exists():
            return observation_from_output(intent, "FAILED", "File not found.", error="File does not exist.")

        try:
            import fitz
        except ImportError:
            return self._fallback(intent, path)

        try:
            doc = fitz.open(str(path))
            text_parts = []
            for page_num in range(min(len(doc), 50)):
                page = doc[page_num]
                text = page.get_text()
                if text.strip():
                    text_parts.append(f"--- Page {page_num + 1} ---\n{text}")
            doc.close()
            content = "\n".join(text_parts)
            truncated = len(content) > 50000
            if truncated:
                content = content[:50000] + "\n\n[Content truncated]"
            return observation_from_output(
                intent, "COMPLETED",
                f"Read PDF: {path} ({len(text_parts)} pages)",
                {"path": str(path), "content": content, "pages": len(text_parts), "truncated": truncated},
            )
        except Exception as exc:
            logger.warning("PDF read error: %s", exc)
            return self._fallback(intent, path)

    def _fallback(self, intent: ToolIntent, path: Path):
        try:
            content = path.read_text(encoding="utf-8", errors="replace")[:10000]
            return observation_from_output(
                intent, "COMPLETED",
                f"Read as raw text: {path}",
                {"path": str(path), "content": content, "method": "raw_text"},
            )
        except Exception as exc:
            return observation_from_output(intent, "FAILED", str(exc), error="PDF read failed.")


class DocxReadTool(Tool):
    def __init__(self, allowed_paths: list[Path]) -> None:
        self.allowed_paths = allowed_paths
        self.definition = ToolDefinition(
            tool_name="document.docx.read",
            display_name="Read Word Document",
            description="Read and extract text from .docx files",
            category="document",
            risk_level=RiskLevel.LOW,
            permissions=["file.read"],
            requires_approval_by_default=False,
        )

    async def execute(self, intent: ToolIntent):
        path = Path(intent.input_payload.get("path", ""))
        if not is_path_allowed(path, self.allowed_paths):
            return observation_from_output(intent, "DENIED", "Path not allowed.", error="Path is not allowed.")
        if not path.exists():
            return observation_from_output(intent, "FAILED", "File not found.", error="File does not exist.")

        try:
            from docx import Document
            doc = Document(str(path))
            paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
            content = "\n".join(paragraphs)
            truncated = len(content) > 50000
            if truncated:
                content = content[:50000] + "\n\n[Content truncated]"
            return observation_from_output(
                intent, "COMPLETED",
                f"Read docx: {path} ({len(paragraphs)} paragraphs)",
                {"path": str(path), "content": content, "paragraphs": len(paragraphs), "truncated": truncated},
            )
        except ImportError:
            return observation_from_output(intent, "FAILED", "python-docx not installed", error="Missing dependency: pip install python-docx")
        except Exception as exc:
            return observation_from_output(intent, "FAILED", str(exc), error="Word read failed.")


class ExcelReadTool(Tool):
    def __init__(self, allowed_paths: list[Path]) -> None:
        self.allowed_paths = allowed_paths
        self.definition = ToolDefinition(
            tool_name="document.xlsx.read",
            display_name="Read Excel",
            description="Read and extract data from Excel files (.xlsx/.xls)",
            category="document",
            risk_level=RiskLevel.LOW,
            permissions=["file.read"],
            requires_approval_by_default=False,
        )

    async def execute(self, intent: ToolIntent):
        path = Path(intent.input_payload.get("path", ""))
        sheet = intent.input_payload.get("sheet", "")
        if not is_path_allowed(path, self.allowed_paths):
            return observation_from_output(intent, "DENIED", "Path not allowed.", error="Path is not allowed.")
        if not path.exists():
            return observation_from_output(intent, "FAILED", "File not found.", error="File does not exist.")

        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
            ws = wb[sheet] if sheet else wb.active
            if not ws:
                return observation_from_output(intent, "FAILED", "Sheet not found", error="Sheet not found.")

            rows = []
            max_rows = min(ws.max_row or 0, 500)
            for row in ws.iter_rows(max_row=max_rows, values_only=True):
                rows.append("\t".join(str(c) if c is not None else "" for c in row))
            wb.close()

            content = "\n".join(rows)
            truncated = len(content) > 50000
            if truncated:
                content = content[:50000] + "\n\n[Content truncated]"
            return observation_from_output(
                intent, "COMPLETED",
                f"Read Excel: {path} ({len(rows)} rows)",
                {"path": str(path), "content": content, "rows": len(rows), "truncated": truncated},
            )
        except ImportError:
            return observation_from_output(intent, "FAILED", "openpyxl not installed", error="Missing dependency: pip install openpyxl")
        except Exception as exc:
            return observation_from_output(intent, "FAILED", str(exc), error="Excel read failed.")
